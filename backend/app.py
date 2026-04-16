"""
SyndicPro Scanner — API Backend v4
Nouveautés : enrichissement Excel asynchrone, cache, context générique.
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from scraper_engine import scrape_all, get_rne_candidates, src_truecaller_query
from scoring_engine import compute_conformity
from db import (init_db, save, get_all, count_all, get_stats, delete_all,
                seed_from_list, set_cache, job_create, job_update, job_get,
                get_result, update_result, invalidate_cache)
from excel_processor import enrich_excel
import os
import io
import csv
import base64
import uuid
import threading
import logging
import openpyxl
import traceback

app    = Flask(__name__, static_folder='../frontend', static_url_path='')
CORS(app)
logging.basicConfig(level=logging.INFO)

init_db()

API_KEY = os.environ.get("API_KEY", "")


def check_key():
    if not API_KEY:
        return True
    return request.headers.get("X-Api-Key") == API_KEY


# ── Recherche ─────────────────────────────────────────────────────────────────

@app.route("/scrape", methods=["POST"])
def scrape():
    if not check_key():
        return jsonify({"error": "Clé API invalide"}), 401

    body    = request.get_json(silent=True) or {}
    name    = (body.get("name")    or "").strip()
    city    = (body.get("city")    or "").strip()
    rne_id  = (body.get("rne_id")  or "").strip()
    context = (body.get("context") or "").strip()

    if not name or not city:
        return jsonify({"error": "Les champs 'name' et 'city' sont obligatoires"}), 400

    try:
        raw        = scrape_all(name, city, rne_id=rne_id, context=context)
        from_cache = len(raw) == 1 and raw[0].get("from_cache", False)
        result     = compute_conformity(raw)
        result["name"] = name
        result["city"] = city
        if from_cache:
            result["from_cache"] = True
        save(result)
        # Mettre en cache si des contacts ont été trouvés
        if result.get("found") or result.get("president"):
            set_cache(name, city, result)
        return jsonify(result)
    except Exception as e:
        app.logger.error(f"Erreur scrape({name},{city}): {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


# ── Candidats RNE (sélection manuelle) ────────────────────────────────────────

@app.route("/rne/candidates")
def rne_candidates():
    name = (request.args.get("name") or "").strip()
    city = (request.args.get("city") or "").strip()
    if not name or not city:
        return jsonify({"error": "Paramètres 'name' et 'city' requis"}), 400
    try:
        candidates = get_rne_candidates(name, city)
        return jsonify({"candidates": candidates})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Stats ──────────────────────────────────────────────────────────────────────

@app.route("/stats")
def stats():
    try:
        return jsonify(get_stats())
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Résultats ──────────────────────────────────────────────────────────────────

@app.route("/results")
def results():
    try:
        limit      = min(int(request.args.get("limit", 200)), 500)
        offset     = int(request.args.get("offset", 0))
        only_found = request.args.get("found", "0") == "1"
        rows  = get_all(limit=limit, offset=offset, only_found=only_found)
        total = count_all(only_found=only_found)
        return jsonify({"results": rows, "count": total})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Modifier un résultat manuellement ─────────────────────────────────────────

@app.route("/results/<int:row_id>", methods=["PUT"])
def update_result_route(row_id):
    if not check_key():
        return jsonify({"error": "Clé API invalide"}), 401
    body = request.get_json(silent=True) or {}
    update_result(row_id, **body)
    return jsonify({"status": "ok"})


@app.route("/results/<int:row_id>", methods=["GET"])
def get_result_route(row_id):
    row = get_result(row_id)
    if not row:
        return jsonify({"error": "Résultat introuvable"}), 404
    return jsonify(row)


@app.route("/results/<int:row_id>/rescrape", methods=["POST"])
def rescrape_result(row_id):
    if not check_key():
        return jsonify({"error": "Clé API invalide"}), 401
    row = get_result(row_id)
    if not row:
        return jsonify({"error": "Résultat introuvable"}), 404
    name   = row["name"]
    city   = row["city"]
    rne_id = row.get("rne_id", "")
    invalidate_cache(name, city)
    try:
        raw    = scrape_all(name, city, rne_id=rne_id, context="")
        result = compute_conformity(raw)
        result["name"] = name
        result["city"] = city
        save(result)
        if result.get("found") or result.get("president"):
            set_cache(name, city, result)
        return jsonify(result)
    except Exception as e:
        app.logger.error(f"Erreur rescrape({name},{city}): {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


# ── Export CSV ─────────────────────────────────────────────────────────────────

@app.route("/export/csv")
def export_csv():
    try:
        rows = get_all(limit=5000)
        out  = io.StringIO()
        w    = csv.DictWriter(out, fieldnames=[
            "id","name","city","phone","email","website",
            "all_phones","all_emails","confidence","sources_hit",
            "president","address","created_at"
        ], extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
        out.seek(0)
        return send_file(
            io.BytesIO(out.getvalue().encode("utf-8-sig")),
            mimetype="text/csv",
            as_attachment=True,
            download_name="contacts.csv"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Export Excel ──────────────────────────────────────────────────────────────

@app.route("/export/excel")
def export_excel_db():
    try:
        only_found = request.args.get("found", "0") == "1"
        rows = get_all(limit=5000, only_found=only_found)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"

        headers = ["ID", "Nom Résidence", "Ville", "Téléphone", "Email", "Site Web",
                   "Président / Gérant", "Adresse", "Tous les tél.", "Tous les emails",
                   "Confiance (%)", "Sources", "Vérifié", "Notes", "Date"]
        ws.append(headers)

        # Style entête
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1E40AF")
            cell.alignment = Alignment(horizontal="center")

        for r in rows:
            members_str = ""
            if r.get("members"):
                members_str = ", ".join(
                    f"{m.get('nom','')} ({m.get('qualite','')})"
                    for m in r["members"]
                )
            ws.append([
                r.get("id", ""),
                r.get("name", ""),
                r.get("city", ""),
                r.get("phone", ""),
                r.get("email", ""),
                r.get("website", ""),
                members_str or r.get("president", ""),
                r.get("address", ""),
                r.get("all_phones", ""),
                r.get("all_emails", ""),
                r.get("confidence", 0),
                r.get("sources_hit", ""),
                "Oui" if r.get("verified") else "",
                r.get("notes", ""),
                r.get("created_at", "")[:16] if r.get("created_at") else "",
            ])

        # Ajuster largeur colonnes
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="contacts_syndicpro.xlsx"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Enrichissement Excel ASYNCHRONE ───────────────────────────────────────────

def _run_excel_job(job_id: str, file_bytes: bytes, context: str):
    """Traitement en arrière-plan — s'exécute dans un thread daemon."""
    try:
        def progress(cur, total):
            job_update(job_id, status="running", progress=cur, total=total)

        result_bytes = enrich_excel(
            io.BytesIO(file_bytes),
            progress_callback=progress,
            context=context
        )
        b64 = base64.b64encode(result_bytes).decode()
        job_update(job_id, status="done", result_b64=b64)
    except Exception as e:
        job_update(job_id, status="error", error=str(e))


@app.route("/enrich/start", methods=["POST"])
def enrich_start():
    if not check_key():
        return jsonify({"error": "Clé API invalide"}), 401
    if "file" not in request.files:
        return jsonify({"error": "Fichier Excel manquant"}), 400

    uploaded = request.files["file"]
    if not uploaded.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Format non supporté (.xlsx / .xls requis)"}), 400

    context    = (request.form.get("context") or "").strip()
    file_bytes = uploaded.read()
    job_id     = str(uuid.uuid4())
    job_create(job_id)

    t = threading.Thread(target=_run_excel_job,
                         args=(job_id, file_bytes, context),
                         daemon=True)
    t.start()
    return jsonify({"job_id": job_id})


@app.route("/enrich/status/<job_id>")
def enrich_status(job_id):
    job = job_get(job_id)
    if not job:
        return jsonify({"error": "Job inconnu"}), 404
    # Ne pas renvoyer le fichier dans le status
    return jsonify({
        "status":   job["status"],
        "progress": job["progress"],
        "total":    job["total"],
        "error":    job.get("error", ""),
    })


@app.route("/enrich/download/<job_id>")
def enrich_download(job_id):
    job = job_get(job_id)
    if not job:
        return jsonify({"error": "Job inconnu"}), 404
    if job["status"] != "done":
        return jsonify({"error": "Fichier pas encore prêt", "status": job["status"]}), 202
    try:
        result_bytes = base64.b64decode(job["result_b64"])
        return send_file(
            io.BytesIO(result_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="contacts_enrichis.xlsx"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Ancienne route synchrone conservée pour compatibilité (max ~5 lignes)
@app.route("/enrich", methods=["POST"])
def enrich():
    if not check_key():
        return jsonify({"error": "Clé API invalide"}), 401
    if "file" not in request.files:
        return jsonify({"error": "Fichier Excel manquant"}), 400
    uploaded = request.files["file"]
    if not uploaded.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Format non supporté"}), 400
    try:
        result_bytes = enrich_excel(uploaded)
        return send_file(
            io.BytesIO(result_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="contacts_enrichis.xlsx"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Import seed RNE ────────────────────────────────────────────────────────────

@app.route("/import/seed", methods=["POST"])
def import_seed():
    if "file" not in request.files:
        return jsonify({"error": "Fichier Excel manquant"}), 400
    uploaded = request.files["file"]
    try:
        wb = openpyxl.load_workbook(uploaded)
        ws = wb.active

        header_row = None
        headers    = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            if any(str(v or '').strip() for v in row):
                header_row = i
                headers    = [str(v or '').strip() for v in row]
                break

        if not header_row:
            return jsonify({"error": "Entêtes non trouvées"}), 400

        def col(frags):
            for frag in frags:
                for i, h in enumerate(headers):
                    if frag.lower() in h.lower():
                        return i
            return None

        name_col = col(["Nom Résidence", "Nom Residence", "Nom", "Dénomination"])
        city_col = col(["Ville", "City"])
        gov_col  = col(["Gouvernorat"])
        rne_col  = col(["ID RNE", "RNE", "Identifiant"])

        if name_col is None:
            return jsonify({"error": "Colonne 'Nom' introuvable"}), 400

        syndics = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            name = str(row[name_col] or "").strip()
            city = str(row[city_col] if city_col is not None else "").strip() or \
                   str(row[gov_col]  if gov_col  is not None else "").strip()
            rne  = str(row[rne_col] if rne_col is not None else "").strip()
            if name and city:
                syndics.append({"name": name, "city": city, "rne_id": rne})

        inserted = seed_from_list(syndics)
        return jsonify({"status": "ok", "inserted": inserted, "total": len(syndics)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Admin ──────────────────────────────────────────────────────────────────────

@app.route("/results/clear", methods=["POST"])
def clear_results():
    if not check_key():
        return jsonify({"error": "Clé API invalide"}), 401
    delete_all()
    return jsonify({"status": "ok"})


@app.route("/health")
def health():
    return jsonify({"status": "ok", "version": "4.0"})


# ── Diagnostic RNE ────────────────────────────────────────────────────────────

@app.route("/debug/rne")
def debug_rne():
    """Teste la connexion RNE et retourne un rapport détaillé."""
    import os, requests as _req
    from scraper_engine import _get_rne_token

    rne_username  = os.environ.get("RNE_USERNAME", "")
    rne_password  = os.environ.get("RNE_PASSWORD", "")
    rne_token_env = os.environ.get("RNE_TOKEN", "")

    report = {
        "env": {
            "RNE_USERNAME": rne_username[:4] + "***" if rne_username else "(non défini)",
            "RNE_PASSWORD": "***" if rne_password else "(non défini)",
            "RNE_TOKEN":    rne_token_env[:8] + "…" if rne_token_env else "(non défini)",
        },
        "token_obtenu": False,
        "token_preview": "",
        "erreur": "",
    }

    try:
        token = _get_rne_token()
        if not token:
            report["erreur"] = "Token vide — vérifiez RNE_USERNAME et RNE_PASSWORD"
            return jsonify(report)
        report["token_obtenu"] = True
        report["token_preview"] = token[:8] + "…"
    except Exception as e:
        report["erreur"] = f"_get_rne_token() : {e}"
        return jsonify(report)

    hdrs = {
        "Authorization": f"Bearer {token}",
        "Referer": "https://www.registre-entreprises.tn/",
        "Accept": "application/json",
    }

    # Étape 1 : trouver un vrai rne_id via l'API PUBLIQUE (sans token, param correct)
    rne_id_test = request.args.get("rne_id", "").strip()
    if not rne_id_test:
        try:
            rs = _req.get(
                "https://www.registre-entreprises.tn/api/rne-api/front-office/shortEntites",
                params={"denominationLatin": "syndic", "size": 5},
                headers={"Referer": "https://www.registre-entreprises.tn/",
                         "Accept": "application/json",
                         "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124.0.0.0 Safari/537.36"},
                timeout=15
            )
            report["search_status"] = rs.status_code
            report["search_raw_preview"] = rs.text[:300]
            if rs.status_code == 200 and rs.text.strip():
                body = rs.json()
                # La réponse est {"registres": [...], "nombreTotal": N}
                items = body.get("registres") or (body if isinstance(body, list) else [])
                if items:
                    rne_id_test = items[0].get("identifiantUnique", "")
                    report["search_sample"] = str(items[0])[:400]
        except Exception as e:
            report["erreur_search"] = str(e)

    report["test_rne_id"] = rne_id_test or "(aucun trouvé)"

    if not rne_id_test:
        report["erreur"] = "Aucun rne_id valide trouvé. Ajoutez ?rne_id=XXXX dans l'URL."
        return jsonify(report)

    # Étape 2 : appel /entites/{id} et affichage brut
    try:
        r = _req.get(
            f"https://www.registre-entreprises.tn/api/rne-api/front-office/entites/{rne_id_test}",
            headers=hdrs, timeout=15
        )
        report["http_status"] = r.status_code
        if r.status_code == 200 and r.text:
            raw = r.json()
            report["champs_disponibles"] = {
                k: (str(v)[:100] if v not in (None, "", [], {}) else "(vide)")
                for k, v in raw.items()
            }
            report["champs_contact"] = {
                k: str(v)
                for k, v in raw.items()
                if any(x in k.lower() for x in ["email","mail","tel","gsm","phone","contact","adresse"])
            }
        else:
            report["erreur"] = f"HTTP {r.status_code} — {r.text[:300]}"
    except Exception as e:
        report["erreur"] = f"Erreur appel /entites : {e}"

    return jsonify(report)


# ── Test scrape complet (debug) ───────────────────────────────────────────────

@app.route("/debug/scrape")
def debug_scrape():
    """
    Lance scrape_all + compute_conformity sur un syndic de test et retourne
    le détail de chaque source (ce qu'elle a trouvé ou non).
    Paramètres : ?name=...&city=...&rne_id=... (rne_id optionnel)
    """
    name   = (request.args.get("name")   or "SYNDIC DES COPROPRIETAIRES DE LA RESIDENCE EL YASSAMINE").strip()
    city   = (request.args.get("city")   or "Sousse").strip()
    rne_id = (request.args.get("rne_id") or "").strip()

    from scraper_engine import scrape_all
    from scoring_engine import compute_conformity

    try:
        raw    = scrape_all(name, city, rne_id=rne_id, context="syndic")
        result = compute_conformity(raw)

        detail = []
        for r in raw:
            detail.append({
                "source":  r.get("source", "?"),
                "phones":  r.get("phones", []),
                "emails":  r.get("emails", []),
                "rne_id":  r.get("rne_id_found", ""),
                "president": r.get("president", ""),
            })

        return jsonify({
            "input":        {"name": name, "city": city, "rne_id": rne_id},
            "sources_brut": detail,
            "resultat":     {
                "phone":       result.get("phone"),
                "email":       result.get("email"),
                "global_conf": result.get("global_conf"),
                "sources_hit": result.get("sources_hit"),
                "found":       result.get("found"),
            }
        })
    except Exception as e:
        import traceback
        return jsonify({"erreur": str(e), "trace": traceback.format_exc()}), 500


# ── Truecaller — auth + test ──────────────────────────────────────────────────

@app.route("/admin/truecaller-otp", methods=["POST"])
def truecaller_otp():
    """
    Étape 1 : demande d'envoi OTP Truecaller sur le numéro de téléphone fourni.
    Body JSON : {"phone": "+21699xxxxxx"}
    Truecaller envoie un SMS avec un code à 6 chiffres.
    """
    import requests as _req
    body  = request.get_json(silent=True) or {}
    phone = (body.get("phone") or "").strip()
    if not phone:
        return jsonify({"error": "Champ 'phone' manquant"}), 400

    # Numéro au format E.164
    if not phone.startswith("+"):
        phone = "+216" + phone.lstrip("0")

    try:
        r = _req.post(
            "https://account.truecaller.com/api/v1/registration/sendOtp",
            json={"phoneNumber": phone, "countryCode": "TN"},
            headers={
                "Content-Type":  "application/json",
                "User-Agent":    "Truecaller/11.75.5 (Android)",
                "clientId":      "4",
            },
            timeout=10
        )
        return jsonify({"status": r.status_code, "body": r.text[:400]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/admin/truecaller-verify", methods=["POST"])
def truecaller_verify():
    """
    Étape 2 : vérification OTP → retourne le Bearer token.
    Body JSON : {"phone": "+21699xxxxxx", "otp": "123456"}
    Copiez le token retourné dans la variable TRUECALLER_TOKEN sur Render.
    """
    import requests as _req
    body  = request.get_json(silent=True) or {}
    phone = (body.get("phone") or "").strip()
    otp   = (body.get("otp")   or "").strip()
    if not phone or not otp:
        return jsonify({"error": "Champs 'phone' et 'otp' requis"}), 400

    if not phone.startswith("+"):
        phone = "+216" + phone.lstrip("0")

    try:
        r = _req.post(
            "https://account.truecaller.com/api/v1/registration/verifyOtp",
            json={"phoneNumber": phone, "countryCode": "TN", "otp": otp},
            headers={
                "Content-Type":  "application/json",
                "User-Agent":    "Truecaller/11.75.5 (Android)",
                "clientId":      "4",
            },
            timeout=10
        )
        body_json = {}
        try:
            body_json = r.json()
        except Exception:
            pass

        token = (body_json.get("token") or body_json.get("access_token") or
                 body_json.get("accessToken") or "")
        return jsonify({
            "status":    r.status_code,
            "token":     token,
            "raw":       r.text[:600],
            "action":    "Copiez 'token' dans la variable TRUECALLER_TOKEN sur Render" if token else "",
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/admin/truecaller-test")
def truecaller_test():
    """
    Teste le token Truecaller stocké dans TRUECALLER_TOKEN.
    Paramètre : ?q=nom (ex: ?q=Lassad+Zitouni)
    """
    from scraper_engine import src_truecaller_query
    token = os.environ.get("TRUECALLER_TOKEN", "").strip()
    q = (request.args.get("q") or "Lassad Zitouni").strip()
    if not token:
        return jsonify({
            "error":   "Variable TRUECALLER_TOKEN non définie",
            "action":  "1) POST /admin/truecaller-otp  2) POST /admin/truecaller-verify  3) Coller le token dans Render"
        }), 400
    result = src_truecaller_query(q)
    return jsonify({
        "query":        q,
        "token_preview": token[:8] + "…",
        "phones":       result.get("phones", []),
        "found":        bool(result.get("phones")),
    })


# ── Pages frontend ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return app.send_static_file("index.html")


@app.route("/dashboard")
def dashboard_page():
    return app.send_static_file("dashboard.html")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port, debug=False)
