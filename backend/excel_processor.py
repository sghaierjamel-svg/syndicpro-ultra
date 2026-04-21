"""
Enrichissement Excel — v5
Améliorations majeures :
  - Mode rapide RNE : quand ID RNE disponible, appel direct RNE (3-5s) avant full scraping
  - Fallback automatique vers scrape_all si RNE ne donne rien
  - Traitement parallèle (5 workers RNE-only, 3 workers full scraping)
  - Colonne Statut (✅ / ❌ / ⚠️)
  - Skip automatique des lignes déjà enrichies
  - Sauvegarde intermédiaire tous les 10 traitements
"""

import io
import time
import json
import logging
import threading
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import PatternFill, Font, Alignment
from scraper_engine import scrape_all, scrape_rne_only
from scoring_engine import compute_conformity

logger = logging.getLogger("excel_processor")

# ── Couleurs ──────────────────────────────────────────────────────────────────
COLOR_HEADER  = "1E3A5F"   # bleu foncé
COLOR_FOUND   = "E8F5E9"   # vert clair
COLOR_NOTFOUND= "FFF8E1"   # orange très clair
COLOR_ERROR   = "FFEBEE"   # rouge très clair
COLOR_SKIP    = "F5F5F5"   # gris clair

PARALLEL_WORKERS     = 3      # workers full scraping
PARALLEL_WORKERS_RNE = 6      # workers mode rapide RNE (endpoints légers)
SAVE_EVERY       = 10         # sauvegarde intermédiaire toutes les N lignes
SLEEP_BETWEEN    = 0.5        # secondes entre chaque ligne (politesse serveurs)


def _style_header(cell):
    cell.font      = Font(bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _color_row(ws, row_num, max_col, hex_color):
    fill = PatternFill("solid", fgColor=hex_color)
    for c in range(1, max_col + 1):
        ws.cell(row_num, c).fill = fill


def enrich_excel(file_obj, progress_callback=None, context=""):
    """
    Enrichit un fichier Excel et retourne les bytes du fichier résultat.

    progress_callback(current, total) — appelé après chaque ligne traitée.
    """
    wb = openpyxl.load_workbook(file_obj)
    ws = wb.active

    # ── Cartographie des entêtes ──────────────────────────────────────────────
    headers    = {}
    header_row = 1
    for ri in range(1, 6):
        row_vals = [str(ws.cell(ri, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if any(v for v in row_vals):
            for ci, v in enumerate(row_vals, 1):
                if v:
                    headers[v] = ci
            if headers:
                header_row = ri
                break

    col_lock = threading.Lock()  # protège les écritures dans ws

    def col_idx(fragments):
        for frag in fragments:
            for h, i in headers.items():
                if frag.lower() in h.lower():
                    return i
        return None

    def ensure_col(label):
        with col_lock:
            if label in headers:
                return headers[label]
            idx = ws.max_column + 1
            cell = ws.cell(header_row, idx, label)
            _style_header(cell)
            headers[label] = idx
            return idx

    # ── Colonnes sources ──────────────────────────────────────────────────────
    name_col = col_idx(["Nom Résidence", "Nom Residence", "Dénomination", "Denomination", "Nom"]) or 1
    city_col = col_idx(["Ville", "Gouvernorat", "City"]) or 2
    ctx_col  = col_idx(["Type", "Activité", "Activite", "Context"])
    rne_col  = col_idx(["ID RNE", "RNE", "Identifiant", "rne_id"])

    # ── Colonnes résultat (créées si absentes) ────────────────────────────────
    status_col   = ensure_col("Statut")
    phone_col    = ensure_col("Téléphone")
    email_col    = ensure_col("Email")
    web_col      = ensure_col("Site Web")
    pres_col     = ensure_col("Président / Gérant")
    members_col  = ensure_col("Membres (bureau)")
    address_col  = ensure_col("Adresse officielle")
    conf_col     = ensure_col("Confiance (%)")
    src_col      = ensure_col("Sources")
    phones_col   = ensure_col("Tous les téléphones")
    emails_col   = ensure_col("Tous les emails")
    rne_out_col  = ensure_col("ID RNE trouvé")
    dur_col      = ensure_col("Durée (s)")

    # ── Lignes à traiter ──────────────────────────────────────────────────────
    data_rows = []
    for ri in range(header_row + 1, ws.max_row + 1):
        name = str(ws.cell(ri, name_col).value or "").strip()
        city = str(ws.cell(ri, city_col).value or "").strip()
        if not name or not city:
            continue

        # Skip si déjà enrichi (téléphone OU email renseigné)
        existing_phone = str(ws.cell(ri, phone_col).value or "").strip()
        existing_email = str(ws.cell(ri, email_col).value or "").strip()
        if existing_phone or existing_email:
            logger.info(f"[Skip] {name} ({city}) — déjà enrichi")
            continue

        data_rows.append(ri)

    total = len(data_rows)
    if total == 0:
        logger.info("[Excel] Aucune ligne à enrichir (toutes déjà remplies).")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    logger.info(f"[Excel] {total} lignes à enrichir avec {PARALLEL_WORKERS} workers parallèles")

    # ── Traitement en parallèle ───────────────────────────────────────────────
    done_count  = 0
    save_counter = 0
    done_lock   = threading.Lock()

    # Séparer les lignes avec/sans ID RNE pour optimiser le traitement
    rows_with_rne    = [r for r in data_rows if rne_col and str(ws.cell(r, rne_col).value or "").strip()]
    rows_without_rne = [r for r in data_rows if r not in rows_with_rne]
    has_rne_batch    = bool(rows_with_rne)

    if has_rne_batch:
        logger.info(f"[Excel] {len(rows_with_rne)} lignes avec ID RNE (mode rapide) + {len(rows_without_rne)} sans ID")

    def process_row(row_num, rne_fast=False):
        """Traite une ligne et retourne un dict de résultats."""
        name   = str(ws.cell(row_num, name_col).value or "").strip()
        city   = str(ws.cell(row_num, city_col).value or "").strip()
        ctx    = str(ws.cell(row_num, ctx_col).value or "").strip() if ctx_col else context
        rne_id = str(ws.cell(row_num, rne_col).value or "").strip() if rne_col else ""
        t0     = time.time()

        try:
            if rne_fast and rne_id:
                # Mode rapide : RNE seul en premier (~3-6s)
                raw  = scrape_rne_only(rne_id, name, city)
                data = compute_conformity(raw)

                # Si RNE ne donne rien d'utile, fallback vers scrape_all complet
                if not data.get("phone") and not data.get("email"):
                    logger.info(f"[RNE fast] Rien trouvé via RNE → fallback full scraping pour {name}")
                    raw  = scrape_all(name, city, rne_id=rne_id, context=ctx)
                    data = compute_conformity(raw)
            else:
                raw  = scrape_all(name, city, rne_id=rne_id, context=ctx)
                data = compute_conformity(raw)

            # Si toujours rien et contexte actif : retry sans contexte
            if not data.get("found") and not data.get("president") and ctx:
                logger.info(f"[Retry] {name} ({city}) — retry sans contexte")
                raw  = scrape_all(name, city, rne_id=rne_id, context="")
                data = compute_conformity(raw)

            dur = round(time.time() - t0, 1)
            return {"row": row_num, "name": name, "city": city, "data": data, "dur": dur, "ok": True}

        except Exception as e:
            dur = round(time.time() - t0, 1)
            logger.error(f"[Excel] Erreur {name} ({city}): {e}", exc_info=True)
            return {"row": row_num, "name": name, "city": city, "data": None,
                    "error": str(e), "dur": dur, "ok": False}

    def write_result(res):
        """Écrit le résultat dans la feuille (thread-safe via col_lock)."""
        row_num = res["row"]
        dur     = res["dur"]

        if not res["ok"]:
            with col_lock:
                ws.cell(row_num, status_col, "⚠️ Erreur")
                ws.cell(row_num, src_col,   f"Erreur: {res.get('error','')[:80]}")
                ws.cell(row_num, dur_col,   dur)
                _color_row(ws, row_num, ws.max_column, COLOR_ERROR)
            return

        data = res["data"]
        members_str = " | ".join(
            f"{m['qualite']}: {m['nom']}"
            for m in data.get("members", [])
        )
        found  = bool(data.get("found") or data.get("president") or data.get("phone") or data.get("email"))
        status = "✅ Trouvé" if found else "❌ Non trouvé"
        color  = COLOR_FOUND if found else COLOR_NOTFOUND

        if found:
            logger.info(f"[Excel] ✅ {res['name']} ({res['city']}) — tél: {data.get('phone','')} email: {data.get('email','')}")
        else:
            logger.info(f"[Excel] ❌ {res['name']} ({res['city']}) — rien trouvé")

        with col_lock:
            ws.cell(row_num, status_col,  status)
            ws.cell(row_num, phone_col,   data.get("phone",    ""))
            ws.cell(row_num, email_col,   data.get("email",    ""))
            ws.cell(row_num, web_col,     data.get("website",  ""))
            ws.cell(row_num, pres_col,    data.get("president",""))
            ws.cell(row_num, members_col, members_str)
            ws.cell(row_num, address_col, data.get("address",  ""))
            ws.cell(row_num, conf_col,    data.get("global_conf", 0))
            ws.cell(row_num, src_col,     ", ".join(data.get("sources_hit", [])))
            ws.cell(row_num, phones_col,  ", ".join(data.get("all_phones",  [])))
            ws.cell(row_num, emails_col,  ", ".join(data.get("all_emails",  [])))
            ws.cell(row_num, rne_out_col, data.get("rne_id", ""))
            ws.cell(row_num, dur_col,     dur)
            _color_row(ws, row_num, ws.max_column, color)

    def _run_batch(row_list, workers, rne_fast=False):
        nonlocal done_count, save_counter
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(process_row, rn, rne_fast): rn for rn in row_list}
            for future in as_completed(futures):
                res = future.result()
                write_result(res)
                with done_lock:
                    done_count  += 1
                    save_counter += 1
                    current = done_count
                if progress_callback:
                    progress_callback(current, total)
                if save_counter >= SAVE_EVERY:
                    save_counter = 0
                    logger.info(f"[Excel] Sauvegarde intermédiaire ({current}/{total})")
                time.sleep(SLEEP_BETWEEN)

    # Traitement : d'abord mode rapide RNE, ensuite lignes sans RNE
    if rows_with_rne:
        logger.info(f"[Excel] Passe 1 — mode rapide RNE ({PARALLEL_WORKERS_RNE} workers)")
        _run_batch(rows_with_rne, PARALLEL_WORKERS_RNE, rne_fast=True)

    if rows_without_rne:
        logger.info(f"[Excel] Passe 2 — full scraping ({PARALLEL_WORKERS} workers)")
        _run_batch(rows_without_rne, PARALLEL_WORKERS, rne_fast=False)

    # ── Ajustement largeur colonnes ───────────────────────────────────────────
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

    # ── Figer la première ligne ───────────────────────────────────────────────
    ws.freeze_panes = ws.cell(header_row + 1, 1)

    # ── Résumé final ──────────────────────────────────────────────────────────
    found_count = sum(
        1 for rn in data_rows
        if str(ws.cell(rn, status_col).value or "").startswith("✅")
    )
    logger.info(
        f"[Excel] Terminé — {found_count}/{total} trouvés "
        f"({round(found_count/total*100) if total else 0}%)"
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
